"""
DPDP Shield — Export API Routes
PDF and Excel export for scan results.
Professional executive compliance report with violations, AI summary, and recommendations.
"""
import io
import json
from fastapi import APIRouter, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable
)
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter

from database import get_db
from models.document import Document
from models.pii_entity import PIIEntity
from compliance.audit_logger import AuditLogger

router = APIRouter()
_audit_logger = AuditLogger()

# ─── Color / Label Constants ─────────────────────────────────────────────────

RISK_HEX = {
    "LOW":      "10b981",
    "MEDIUM":   "f59e0b",
    "HIGH":     "f97316",
    "CRITICAL": "ef4444",
}

DPDP_SECTION_DETAILS = {
    "4":  ("Section 4 — Unlawful Purpose",     "Data collected without a defined lawful purpose."),
    "6":  ("Section 6 — Consent",               "Explicit, informed consent not obtained or recorded."),
    "8":  ("Section 8 — Fiduciary Duty",        "Sensitive data unmasked or inadequately secured."),
    "9":  ("Section 9 — Children's Data",       "Age-sensitive PII detected without parental safeguards."),
    "16": ("Section 16 — Cross-border Transfer","Foreign personal data transferred without notification."),
}

DOC_TYPE_LABELS = {
    "hr_record":      "HR Record",
    "customer_kyc":   "Customer KYC",
    "legal_contract": "Legal Contract",
    "invoice":        "Financial Invoice",
    "marketing":      "Marketing List",
    "medical":        "Medical Record",
    "resume":         "Resume / CV",
    "general":        "General Document",
}

REMEDIATION_BY_SECTION = {
    "6":  "Implement explicit consent collection with opt-in mechanisms at all data entry points.",
    "8":  "Apply data masking (Aadhaar: XXXX XXXX 1012, PAN: ABCXXXXXXF) and encrypt sensitive fields.",
    "9":  "Add age-verification checks and obtain verifiable parental consent before processing.",
    "16": "Register cross-border data transfers with the Data Protection Board as required.",
    "4":  "Define a lawful purpose statement and record it for every data collection operation.",
}


# ─── API Endpoints ────────────────────────────────────────────────────────────

@router.get("/{document_id}/pdf")
async def export_pdf(
    document_id: str,
    request: Request = None,
    db: AsyncSession = Depends(get_db),
):
    """Export compliance report as a downloadable PDF."""
    doc = await db.get(Document, document_id)
    if not doc:
        raise HTTPException(404, "Document not found")

    result = await db.execute(
        select(PIIEntity).where(PIIEntity.document_id == document_id)
    )
    pii_entities = result.scalars().all()

    buffer = io.BytesIO()
    _build_pdf(buffer, doc, pii_entities)
    buffer.seek(0)

    safe_name = (doc.original_filename or "document").replace(" ", "_")
    date_str  = datetime.now().strftime("%Y%m%d")

    await _audit_logger.log(
        action='EXPORT_PDF',
        document_id=document_id,
        details={'filename': doc.original_filename or doc.filename},
        request=request,
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="dpdp_report_{safe_name}_{date_str}.pdf"'},
    )


@router.get("/{document_id}/excel")
async def export_excel(
    document_id: str,
    request: Request = None,
    db: AsyncSession = Depends(get_db),
):
    """Export PII findings as a multi-sheet Excel workbook."""
    doc = await db.get(Document, document_id)
    if not doc:
        raise HTTPException(404, "Document not found")

    result = await db.execute(
        select(PIIEntity).where(PIIEntity.document_id == document_id)
    )
    pii_entities = result.scalars().all()

    buffer = io.BytesIO()
    _build_excel(buffer, doc, pii_entities)
    buffer.seek(0)

    safe_name = (doc.original_filename or "document").replace(" ", "_")
    date_str  = datetime.now().strftime("%Y%m%d")

    await _audit_logger.log(
        action='EXPORT_EXCEL',
        document_id=document_id,
        details={'filename': doc.original_filename or doc.filename},
        request=request,
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="dpdp_pii_{safe_name}_{date_str}.xlsx"'},
    )


# ─── PDF Builder ──────────────────────────────────────────────────────────────

def _build_pdf(buffer, doc: Document, entities: list):
    """
    Build a professional executive compliance report PDF.
    Sections: Header · Doc Info · Risk Assessment · AI Summary ·
              DPDP Violations · Remediation · PII Findings · Footer
    """
    pdf = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50
    )
    styles   = getSampleStyleSheet()
    navy     = HexColor("#0f172a")
    orange   = HexColor("#EB6A2A")
    light    = HexColor("#f8f5ef")
    white    = HexColor("#ffffff")
    slate    = HexColor("#64748b")
    red_h    = HexColor("#dc2626")
    red_bg   = HexColor("#fee8e8")
    grn_h    = HexColor("#28a745")
    grn_bg   = HexColor("#e8f5ec")
    ai_bg    = HexColor("#fef0e6")
    ai_bdr   = HexColor("#f8c9a8")

    risk_hex   = RISK_HEX.get(doc.risk_level or "LOW", "10b981")
    risk_color = HexColor(f"#{risk_hex}")

    doc_type_label   = DOC_TYPE_LABELS.get(doc.document_type or "general", "General Document")
    compliance_score = max(0, 100 - int(doc.risk_score or 0))

    try:
        violations = json.loads(doc.violations_json or "[]")
    except Exception:
        violations = []

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Paragraph(
        '<font color="#EB6A2A" size="22"><b>DPDP Shield</b></font>',
        ParagraphStyle("hdr", fontSize=22, spaceAfter=2)
    ))
    story.append(Paragraph(
        "Privacy Compliance Report  \u00b7  Digital Personal Data Protection Act 2023",
        ParagraphStyle("sub", fontSize=10, textColor=slate, spaceAfter=14)
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=orange))
    story.append(Spacer(1, 16))

    # ── Executive Summary Row ─────────────────────────────────────────────────
    exec_rows = [
        ["Risk Level", "Risk Score", "Compliance", "Classification"],
        [doc.risk_level or "N/A", f"{doc.risk_score or 0:.0f} / 100",
         f"{compliance_score}%", doc_type_label],
    ]
    exec_tbl = Table(exec_rows, colWidths=[1.55*inch, 1.55*inch, 1.55*inch, 1.75*inch])
    exec_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR",  (0, 0), (-1, 0), white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, 1), (0, 1), risk_color),
        ("TEXTCOLOR",  (0, 1), (0, 1), white),
        ("FONTNAME",   (0, 1), (0, 1), "Helvetica-Bold"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING",    (0, 0), (-1, -1), 10),
    ]))
    story.append(exec_tbl)
    story.append(Spacer(1, 18))

    # ── Document Information ──────────────────────────────────────────────────
    story.append(Paragraph("<b>Document Information</b>", styles["Heading2"]))
    scan_date = doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "N/A"
    info_rows = [
        ["Field", "Value"],
        ["Filename",         doc.original_filename or "N/A"],
        ["File Type",        (doc.file_type or "").upper()],
        ["Classification",   doc_type_label],
        ["Scan Date",        scan_date],
        ["Pages Scanned",    str(doc.page_count or 1)],
        ["File Size",        f"{(doc.file_size or 0) / 1024:.1f} KB"],
        ["Total PII Found",  str(len(entities))],
        ["Retention Policy", doc.retention_policy or "Per company policy"],
        ["Department",       doc.department or "General"],
    ]
    info_tbl = Table(info_rows, colWidths=[2.2*inch, 4*inch])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), navy),
        ("TEXTCOLOR",     (0, 0), (-1, 0), white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 10),
        ("BACKGROUND",    (0, 1), (0, -1), light),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [white, light]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING",       (0, 0), (-1, -1), 8),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 18))

    # ── AI Risk Analysis ──────────────────────────────────────────────────────
    if doc.ai_summary:
        story.append(Paragraph("<b>AI Risk Analysis</b>", styles["Heading2"]))
        ai_tbl = Table(
            [[Paragraph(
                doc.ai_summary,
                ParagraphStyle("ai", fontSize=10, leading=16, textColor=HexColor("#1a1a1a"))
            )]],
            colWidths=[6.2*inch]
        )
        ai_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), ai_bg),
            ("GRID",       (0, 0), (-1, -1), 0.5, ai_bdr),
            ("PADDING",    (0, 0), (-1, -1), 12),
        ]))
        story.append(ai_tbl)
        story.append(Spacer(1, 18))

    # ── DPDP Act 2023 Violations ──────────────────────────────────────────────
    story.append(Paragraph("<b>DPDP Act 2023 \u2014 Violations Detected</b>", styles["Heading2"]))
    if violations:
        vio_data = [["Section", "Violation Title", "Description"]]
        for sec in violations:
            title, desc = DPDP_SECTION_DETAILS.get(
                str(sec), (f"Section {sec}", "Compliance violation detected.")
            )
            vio_data.append([f"\u00a7{sec}", title, desc])
        vio_tbl = Table(vio_data, colWidths=[0.6*inch, 2.2*inch, 3.4*inch])
        vio_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), red_h),
            ("TEXTCOLOR",     (0, 0), (-1, 0), white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("BACKGROUND",    (0, 1), (-1, -1), red_bg),
            ("TEXTCOLOR",     (0, 1), (-1, -1), red_h),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING",       (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(vio_tbl)
    else:
        ok_tbl = Table(
            [["\u2713  No DPDP violations detected \u2014 document passes all compliance checks."]],
            colWidths=[6.2*inch]
        )
        ok_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), grn_bg),
            ("TEXTCOLOR",  (0, 0), (-1, -1), grn_h),
            ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
            ("PADDING",    (0, 0), (-1, -1), 10),
            ("GRID",       (0, 0), (-1, -1), 0.5, HexColor("#a8d5b5")),
        ]))
        story.append(ok_tbl)
    story.append(Spacer(1, 18))

    # ── Remediation Recommendations ───────────────────────────────────────────
    if violations:
        story.append(Paragraph("<b>Remediation Recommendations</b>", styles["Heading2"]))
        rem_data = [["Priority", "Action Required"]]
        for sec in violations:
            action   = REMEDIATION_BY_SECTION.get(
                str(sec), f"Review and address Section {sec} compliance requirements."
            )
            priority = "Immediate" if str(sec) in ("8", "6") else "Within 30 days"
            rem_data.append([priority, action])
        rem_data.append([
            "Within 90 days",
            doc.retention_policy or
            "Define and automate data deletion after the processing purpose is served.",
        ])
        rem_tbl = Table(rem_data, colWidths=[1.4*inch, 4.8*inch])
        rem_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), navy),
            ("TEXTCOLOR",     (0, 0), (-1, 0), white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [white, light]),
            ("GRID",          (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("PADDING",       (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(rem_tbl)
        story.append(Spacer(1, 18))

    # ── PII Findings table (max 50 rows) ──────────────────────────────────────
    if entities:
        story.append(Paragraph("<b>PII Entities Detected</b>", styles["Heading2"]))
        if len(entities) > 50:
            story.append(Paragraph(
                f'<font size="9" color="#94a3b8">Showing first 50 of {len(entities)} entities. '
                'Download Excel export for the full list.</font>',
                ParagraphStyle("note", fontSize=9, spaceAfter=6)
            ))
        pii_headers = ["#", "Entity Type", "Masked Value", "Source", "Confidence", "Page"]
        pii_data    = [pii_headers]
        for i, e in enumerate(entities[:50], 1):
            pii_data.append([
                str(i),
                e.entity_type,
                e.masked_value or "***",
                e.detector_source or "\u2014",
                f"{e.confidence:.0%}",
                str(e.page_num or 1),
            ])
        col_w = [0.4*inch, 1.6*inch, 2.1*inch, 1.0*inch, 0.9*inch, 0.5*inch]
        pii_tbl = Table(pii_data, colWidths=col_w)
        pii_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), navy),
            ("TEXTCOLOR",     (0, 0), (-1, 0), white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [white, light]),
            ("GRID",          (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("PADDING",       (0, 0), (-1, -1), 5),
        ]))
        story.append(pii_tbl)

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 30))
    story.append(HRFlowable(width="100%", thickness=1, color=HexColor("#e2e8f0")))
    story.append(Paragraph(
        f'<font size="8" color="#94a3b8">Generated by DPDP Shield AI Privacy Intelligence Platform  \u00b7  '
        f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}  \u00b7  '
        f'Digital Personal Data Protection Act 2023</font>',
        ParagraphStyle("footer", fontSize=8, textColor=HexColor("#94a3b8"), spaceBefore=8),
    ))
    pdf.build(story)


# ─── Excel Builder ────────────────────────────────────────────────────────────

def _build_excel(buffer, doc: Document, entities: list):
    wb        = openpyxl.Workbook()
    H_FILL    = PatternFill("solid", fgColor="0F172A")
    ALT_FILL  = PatternFill("solid", fgColor="F1F5F9")
    H_FONT    = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14, color="EB6A2A")
    risk_hex  = RISK_HEX.get(doc.risk_level or "LOW", "10b981")

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "DPDP Shield \u2014 Compliance Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    scan_date = doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "N/A"
    doc_type_label = DOC_TYPE_LABELS.get(doc.document_type or "general", "General Document")

    try:
        violations = json.loads(doc.violations_json or "[]")
    except Exception:
        violations = []

    rows = [
        ("Filename",          doc.original_filename or "N/A"),
        ("File Type",         (doc.file_type or "").upper()),
        ("Classification",    doc_type_label),
        ("Scan Date",         scan_date),
        ("Risk Level",        doc.risk_level or "N/A"),
        ("Risk Score",        f"{doc.risk_score or 0:.0f} / 100"),
        ("Compliance Score",  f"{max(0, 100 - int(doc.risk_score or 0))}%"),
        ("Total PII Found",   len(entities)),
        ("Pages Scanned",     doc.page_count or 1),
        ("DPDP Violations",   ", ".join(f"\u00a7{s}" for s in violations) or "None"),
        ("Retention Policy",  doc.retention_policy or "Per company policy"),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = str(v)
        ws[f"A{i}"].font = Font(bold=True)
        fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        ws[f"A{i}"].fill = fill
        ws[f"B{i}"].fill = fill
        if k == "Risk Level":
            ws[f"B{i}"].fill = PatternFill("solid", fgColor=risk_hex)
            ws[f"B{i}"].font = Font(bold=True, color="FFFFFF")

    # AI Summary if present
    if doc.ai_summary:
        ai_row = len(rows) + 4
        ws[f"A{ai_row}"] = "AI Risk Summary"
        ws[f"A{ai_row}"].font = Font(bold=True)
        ws[f"B{ai_row}"] = doc.ai_summary
        ws[f"B{ai_row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[ai_row].height = 60

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 52

    # ── Sheet 2: PII Entities ─────────────────────────────────────
    ws2 = wb.create_sheet("PII Entities")
    headers = ["#", "Entity Type", "Original Value", "Masked Value",
               "Confidence", "Source", "Page", "Char Start", "Char End"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = H_FILL
        c.font = H_FONT
        c.alignment = Alignment(horizontal="center")

    for r, e in enumerate(entities, 2):
        row_fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            r - 1, e.entity_type, e.entity_value or "",
            e.masked_value or "", f"{e.confidence:.0%}",
            e.detector_source or "", e.page_num or 1,
            e.start_char or 0, e.end_char or 0,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.fill = row_fill

    for col, width in zip(range(1, 10), [5, 20, 30, 30, 12, 12, 8, 12, 12]):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.auto_filter.ref = f"A1:I{len(entities) + 1}"

    # ── Sheet 3: PII Breakdown ────────────────────────────────────
    ws3 = wb.create_sheet("PII Breakdown")
    for col, h in enumerate(["PII Type", "Count", "% of Total"], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = H_FILL
        c.font = H_FONT
        c.alignment = Alignment(horizontal="center")

    total = len(entities)
    for i, (etype, count) in enumerate(Counter(e.entity_type for e in entities).most_common(), 2):
        fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        ws3.cell(row=i, column=1, value=etype).fill = fill
        ws3.cell(row=i, column=2, value=count).fill = fill
        ws3.cell(row=i, column=3, value=f"{count/total*100:.1f}%" if total else "0%").fill = fill

    for col, w in [(1, 25), (2, 12), (3, 14)]:
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Violations & Remediation ────────────────────────
    if violations:
        ws4 = wb.create_sheet("Violations")
        for col, h in enumerate(["Section", "Violation", "Description", "Remediation", "Priority"], 1):
            c = ws4.cell(row=1, column=col, value=h)
            c.fill = PatternFill("solid", fgColor="DC2626")
            c.font = Font(bold=True, color="FFFFFF", size=10)

        for i, sec in enumerate(violations, 2):
            title, desc = DPDP_SECTION_DETAILS.get(
                str(sec), (f"Section {sec}", "Compliance violation.")
            )
            action   = REMEDIATION_BY_SECTION.get(str(sec), "Review compliance requirements.")
            priority = "Immediate" if str(sec) in ("8", "6") else "Within 30 days"
            fill = PatternFill("solid", fgColor="FEE8E8") if i % 2 == 0 else PatternFill("solid", fgColor="FFF5F5")
            for col, val in enumerate([f"\u00a7{sec}", title, desc, action, priority], 1):
                c = ws4.cell(row=i, column=col, value=val)
                c.fill = fill
                c.alignment = Alignment(wrap_text=True)
        for col, w in [(1, 10), (2, 28), (3, 40), (4, 50), (5, 16)]:
            ws4.column_dimensions[get_column_letter(col)].width = w

    wb.save(buffer)
